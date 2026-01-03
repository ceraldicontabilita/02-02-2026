"""
Invoice Metadata Router
API endpoints per gestione metadata fatture
"""
from fastapi import APIRouter, Depends, Query, Path, status
from typing import List, Dict, Any, Optional
from datetime import datetime
import logging

from app.database import Database, Collections
from app.repositories.invoice_repository import InvoiceRepository
from app.models.invoice import (
    InvoiceMetadataCreate,
    InvoiceMetadataTemplate,
    InvoiceMetadataResponse
)
from app.utils.dependencies import get_current_user
from app.exceptions import NotFoundError, ValidationError

logger = logging.getLogger(__name__)
router = APIRouter()


# ==================== ENDPOINT 27: Get Metadata Fattura ====================
@router.get(
    "/{invoice_id}/metadata",
    summary="Get metadata fattura",
    description="Recupera metadata associati a una fattura"
)
async def get_invoice_metadata(
    invoice_id: str = Path(..., description="ID fattura"),
    current_user: Dict[str, Any] = Depends(get_current_user)
) -> Dict[str, Any]:
    """
    Recupera metadata fattura.
    
    I metadata sono campi personalizzati che possono essere
    associati alle fatture per aggiungere informazioni extra.
    """
    db = Database.get_db()
    invoices_collection = db[Collections.INVOICES]
    
    invoice = await invoices_collection.find_one({'_id': invoice_id})
    if not invoice:
        raise NotFoundError(f"Invoice {invoice_id} not found")
    
    metadata = invoice.get('metadata', {})
    
    return {
        'invoice_id': str(invoice['_id']),
        'invoice_number': invoice.get('invoice_number'),
        'metadata': metadata,
        'has_metadata': len(metadata) > 0
    }


# ==================== ENDPOINT 28: Aggiorna Metadata ====================
@router.put(
    "/{invoice_id}/metadata",
    summary="Aggiorna metadata",
    description="Aggiorna o crea metadata per una fattura"
)
async def update_invoice_metadata(
    invoice_id: str = Path(..., description="ID fattura"),
    metadata: InvoiceMetadataCreate = ...,
    current_user: Dict[str, Any] = Depends(get_current_user)
) -> Dict[str, Any]:
    """
    Aggiorna metadata fattura.
    
    **Utilizzo:**
    - Aggiungi campi personalizzati
    - Traccia informazioni extra
    - Categorizza fatture
    
    **Esempio:**
    ```json
    {
        "invoice_id": "...",
        "metadata_fields": [
            {
                "field_name": "centro_di_costo",
                "field_type": "text",
                "field_value": "Reparto Produzione",
                "required": false
            },
            {
                "field_name": "budget_residuo",
                "field_type": "number",
                "field_value": 5000.00,
                "required": true
            }
        ],
        "template_name": "Fatture Produzione"
    }
    ```
    """
    db = Database.get_db()
    invoices_collection = db[Collections.INVOICES]
    
    invoice = await invoices_collection.find_one({'_id': invoice_id})
    if not invoice:
        raise NotFoundError(f"Invoice {invoice_id} not found")
    
    # Prepara metadata
    metadata_dict = {}
    for field in metadata.metadata_fields:
        metadata_dict[field.field_name] = {
            'type': field.field_type,
            'value': field.field_value,
            'required': field.required
        }
    
    # Aggiorna
    await invoices_collection.update_one(
        {'_id': invoice_id},
        {
            '$set': {
                'metadata': metadata_dict,
                'metadata_template': metadata.template_name,
                'updated_at': datetime.utcnow()
            }
        }
    )
    
    logger.info(f"Updated metadata for invoice {invoice_id}")
    
    return {
        'invoice_id': invoice_id,
        'invoice_number': invoice.get('invoice_number'),
        'metadata': metadata_dict,
        'template_name': metadata.template_name,
        'fields_count': len(metadata.metadata_fields)
    }


# ==================== ENDPOINT 29: Salva Template Metadata ====================
@router.post(
    "/metadata/save-template",
    status_code=status.HTTP_201_CREATED,
    summary="Salva template metadata",
    description="Crea template riutilizzabile per metadata"
)
async def save_metadata_template(
    template: InvoiceMetadataTemplate,
    current_user: Dict[str, Any] = Depends(get_current_user)
) -> Dict[str, Any]:
    """
    Salva template metadata.
    
    I template permettono di riutilizzare la stessa
    struttura di metadata per più fatture.
    
    **Utilità:**
    - Standardizza metadata
    - Risparmia tempo
    - Garantisce consistenza
    """
    db = Database.get_db()
    templates_collection = db[Collections.INVOICE_METADATA_TEMPLATES]
    
    # Verifica che non esista già
    existing = await templates_collection.find_one({
        'template_name': template.template_name
    })
    
    if existing:
        raise ValidationError(f"Template '{template.template_name}' already exists")
    
    # Prepara template
    template_data = {
        'template_name': template.template_name,
        'description': template.description,
        'fields': [field.model_dump() for field in template.fields],
        'apply_to_supplier': template.apply_to_supplier,
        'created_by': current_user['user_id'],
        'created_at': datetime.utcnow()
    }
    
    result = await templates_collection.insert_one(template_data)
    
    logger.info(f"Created metadata template: {template.template_name}")
    
    return {
        'id': str(result.inserted_id),
        'template_name': template.template_name,
        'description': template.description,
        'fields_count': len(template.fields),
        'apply_to_supplier': template.apply_to_supplier
    }


# ==================== ENDPOINT 30: Lista Templates ====================
@router.get(
    "/metadata/templates",
    summary="Lista templates metadata",
    description="Recupera tutti i template metadata salvati"
)
async def list_metadata_templates(
    current_user: Dict[str, Any] = Depends(get_current_user)
) -> List[Dict[str, Any]]:
    """
    Lista templates metadata.
    
    Ritorna tutti i template disponibili per applicare
    metadata alle fatture.
    """
    db = Database.get_db()
    templates_collection = db[Collections.INVOICE_METADATA_TEMPLATES]
    
    cursor = templates_collection.find().sort('template_name', 1)
    templates = await cursor.to_list(length=None)
    
    result = []
    for template in templates:
        result.append({
            'id': str(template['_id']),
            'template_name': template.get('template_name'),
            'description': template.get('description'),
            'fields_count': len(template.get('fields', [])),
            'apply_to_supplier': template.get('apply_to_supplier'),
            'created_at': template.get('created_at')
        })
    
    return result


# ==================== ENDPOINT 31: Elimina Template ====================
@router.delete(
    "/metadata/templates/{template_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="Elimina template",
    description="Elimina template metadata"
)
async def delete_metadata_template(
    template_id: str = Path(..., description="ID template"),
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """Elimina template metadata."""
    db = Database.get_db()
    templates_collection = db[Collections.INVOICE_METADATA_TEMPLATES]
    
    result = await templates_collection.delete_one({'_id': template_id})
    
    if result.deleted_count == 0:
        raise NotFoundError(f"Template {template_id} not found")
    
    logger.info(f"Deleted template {template_id}")
    return None


# ==================== ENDPOINT 32: Applica Template ====================
@router.post(
    "/metadata/apply-template",
    summary="Applica template",
    description="Applica template metadata a una o più fatture"
)
async def apply_metadata_template(
    invoice_ids: List[str] = Query(..., description="IDs fatture"),
    template_name: str = Query(..., description="Nome template"),
    current_user: Dict[str, Any] = Depends(get_current_user)
) -> Dict[str, Any]:
    """
    Applica template a fatture.
    
    Applica lo stesso template metadata a multiple fatture
    contemporaneamente.
    """
    db = Database.get_db()
    templates_collection = db[Collections.INVOICE_METADATA_TEMPLATES]
    invoices_collection = db[Collections.INVOICES]
    
    # Recupera template
    template = await templates_collection.find_one({
        'template_name': template_name
    })
    
    if not template:
        raise NotFoundError(f"Template '{template_name}' not found")
    
    # Prepara metadata da template
    metadata_dict = {}
    for field in template.get('fields', []):
        metadata_dict[field['field_name']] = {
            'type': field['field_type'],
            'value': field.get('field_value'),
            'required': field.get('required', False)
        }
    
    # Applica a tutte le fatture
    updated_count = 0
    for invoice_id in invoice_ids:
        result = await invoices_collection.update_one(
            {'_id': invoice_id},
            {
                '$set': {
                    'metadata': metadata_dict,
                    'metadata_template': template_name,
                    'updated_at': datetime.utcnow()
                }
            }
        )
        if result.modified_count > 0:
            updated_count += 1
    
    logger.info(f"Applied template '{template_name}' to {updated_count} invoices")
    
    return {
        'template_name': template_name,
        'invoices_requested': len(invoice_ids),
        'invoices_updated': updated_count,
        'fields_applied': len(metadata_dict)
    }


# ==================== ENDPOINT 33: Applica Bulk ====================
@router.post(
    "/metadata/apply-bulk",
    summary="Applica metadata bulk",
    description="Applica metadata personalizzati a multiple fatture"
)
async def apply_metadata_bulk(
    metadata_list: List[InvoiceMetadataCreate],
    current_user: Dict[str, Any] = Depends(get_current_user)
) -> Dict[str, Any]:
    """
    Applica metadata bulk.
    
    Permette di applicare metadata diversi a fatture diverse
    in un'unica operazione.
    """
    db = Database.get_db()
    invoices_collection = db[Collections.INVOICES]
    
    updated = 0
    errors = []
    
    for metadata in metadata_list:
        try:
            # Prepara metadata
            metadata_dict = {}
            for field in metadata.metadata_fields:
                metadata_dict[field.field_name] = {
                    'type': field.field_type,
                    'value': field.field_value,
                    'required': field.required
                }
            
            # Aggiorna
            result = await invoices_collection.update_one(
                {'_id': metadata.invoice_id},
                {
                    '$set': {
                        'metadata': metadata_dict,
                        'metadata_template': metadata.template_name,
                        'updated_at': datetime.utcnow()
                    }
                }
            )
            
            if result.modified_count > 0:
                updated += 1
        
        except Exception as e:
            errors.append({
                'invoice_id': metadata.invoice_id,
                'error': str(e)
            })
    
    return {
        'total': len(metadata_list),
        'updated': updated,
        'errors': len(errors),
        'error_details': errors
    }


# ==================== ENDPOINT 34: Lista Campi Disponibili ====================
@router.get(
    "/metadata/fields",
    summary="Lista campi metadata",
    description="Lista tutti i tipi di campi metadata disponibili"
)
async def list_metadata_fields(
    current_user: Dict[str, Any] = Depends(get_current_user)
) -> Dict[str, Any]:
    """
    Lista campi metadata disponibili.
    
    Ritorna i tipi di campi che possono essere usati
    nei metadata delle fatture.
    """
    return {
        'field_types': [
            {
                'type': 'text',
                'description': 'Testo libero',
                'example': 'Centro di costo Produzione'
            },
            {
                'type': 'number',
                'description': 'Numero decimale',
                'example': 1500.50
            },
            {
                'type': 'date',
                'description': 'Data',
                'example': '2024-12-30'
            },
            {
                'type': 'boolean',
                'description': 'Si/No',
                'example': True
            },
            {
                'type': 'select',
                'description': 'Selezione da lista',
                'example': 'Opzione A'
            }
        ],
        'common_fields': [
            'centro_di_costo',
            'budget',
            'responsabile',
            'progetto',
            'commessa',
            'categoria_spesa',
            'priorita',
            'stato_approvazione'
        ]
    }


# ==================== ENDPOINT 35: Crea Campo Custom ====================
@router.post(
    "/metadata/custom-field",
    summary="Crea campo custom",
    description="Definisce nuovo tipo di campo metadata custom"
)
async def create_custom_metadata_field(
    field_name: str = Query(..., description="Nome campo"),
    field_type: str = Query(..., description="Tipo campo"),
    description: Optional[str] = Query(None, description="Descrizione"),
    required: bool = Query(False, description="Obbligatorio"),
    current_user: Dict[str, Any] = Depends(get_current_user)
) -> Dict[str, Any]:
    """
    Crea campo metadata custom.
    
    Definisce un nuovo tipo di campo che può essere
    riutilizzato nei template e nelle fatture.
    """
    db = Database.get_db()
    custom_fields_collection = db['invoice_metadata_custom_fields']
    
    # Verifica unicità
    existing = await custom_fields_collection.find_one({
        'field_name': field_name
    })
    
    if existing:
        raise ValidationError(f"Field '{field_name}' already exists")
    
    # Crea campo
    field_data = {
        'field_name': field_name,
        'field_type': field_type,
        'description': description,
        'required': required,
        'created_by': current_user['user_id'],
        'created_at': datetime.utcnow()
    }
    
    result = await custom_fields_collection.insert_one(field_data)
    
    logger.info(f"Created custom field: {field_name}")
    
    return {
        'id': str(result.inserted_id),
        'field_name': field_name,
        'field_type': field_type,
        'description': description,
        'required': required
    }


# ==================== ENDPOINT 36: Mesi Archiviati ====================
@router.get(
    "/archived-months",
    summary="Mesi archiviati",
    description="Lista mesi con fatture archiviate"
)
async def get_archived_months(
    current_user: Dict[str, Any] = Depends(get_current_user)
) -> List[str]:
    """
    Lista mesi archiviati.
    
    Ritorna tutti i mesi per cui esistono fatture archiviate.
    Utile per organizzare l'archivio.
    """
    db = Database.get_db()
    invoices_collection = db[Collections.INVOICES]
    
    # Trova tutti i month_year distinti per fatture archiviate
    pipeline = [
        {'$match': {'archived': True}},
        {'$group': {'_id': '$month_year'}},
        {'$sort': {'_id': -1}}
    ]
    
    cursor = invoices_collection.aggregate(pipeline)
    months = await cursor.to_list(length=None)
    
    return [month['_id'] for month in months if month.get('_id')]


# ==================== ENDPOINT 37: Archivia Mese ====================
@router.post(
    "/archive-month",
    summary="Archivia mese",
    description="Archivia tutte le fatture di un mese"
)
async def archive_month(
    month_year: str = Query(..., description="Mese anno formato YYYY-MM"),
    current_user: Dict[str, Any] = Depends(get_current_user)
) -> Dict[str, Any]:
    """
    Archivia mese.
    
    Marca come archiviate tutte le fatture del mese specificato.
    Le fatture archiviate sono escluse dalle ricerche normali.
    """
    db = Database.get_db()
    invoices_collection = db[Collections.INVOICES]
    
    result = await invoices_collection.update_many(
        {'month_year': month_year},
        {
            '$set': {
                'archived': True,
                'archived_at': datetime.utcnow(),
                'archived_by': current_user['user_id']
            }
        }
    )
    
    logger.info(f"Archived {result.modified_count} invoices for {month_year}")
    
    return {
        'month_year': month_year,
        'invoices_archived': result.modified_count
    }


# ==================== ENDPOINT 38: Cerca per Metadata ====================
@router.get(
    "/search-by-metadata",
    summary="Cerca per metadata",
    description="Cerca fatture che hanno specifici valori nei metadata"
)
async def search_by_metadata(
    field_name: str = Query(..., description="Nome campo metadata"),
    field_value: str = Query(..., description="Valore da cercare"),
    current_user: Dict[str, Any] = Depends(get_current_user)
) -> List[Dict[str, Any]]:
    """
    Cerca fatture per metadata.
    
    Permette di filtrare fatture in base ai valori
    dei campi metadata personalizzati.
    
    **Esempi:**
    - `field_name=centro_di_costo&field_value=Produzione`
    - `field_name=progetto&field_value=Ristrutturazione`
    """
    db = Database.get_db()
    invoices_collection = db[Collections.INVOICES]
    
    # Query dinamica per metadata
    query = {
        f'metadata.{field_name}.value': field_value
    }
    
    cursor = invoices_collection.find(query).sort('invoice_date', -1)
    invoices = await cursor.to_list(length=100)
    
    result = []
    for invoice in invoices:
        result.append({
            'id': str(invoice['_id']),
            'invoice_number': invoice.get('invoice_number'),
            'invoice_date': invoice.get('invoice_date'),
            'supplier_name': invoice.get('supplier_name'),
            'total_amount': invoice.get('total_amount'),
            'metadata': invoice.get('metadata', {})
        })
    
    return result


# ==================== ENDPOINT 39: Export Metadata ====================
@router.get(
    "/metadata/export",
    summary="Export metadata Excel",
    description="Export report con tutte le fatture e i loro metadata"
)
async def export_metadata_report(
    month_year: Optional[str] = Query(None, description="Filtra per mese"),
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """
    Export metadata report.
    
    Genera Excel con tutte le fatture e relativi metadata
    per analisi e reportistica.
    """
    db = Database.get_db()
    invoices_collection = db[Collections.INVOICES]
    
    query = {}
    if month_year:
        query['month_year'] = month_year
    
    cursor = invoices_collection.find(query).sort('invoice_date', -1)
    invoices = await cursor.to_list(length=1000)
    
    # Prepara dati per Excel
    rows = []
    for invoice in invoices:
        metadata = invoice.get('metadata', {})
        
        row = {
            'Numero Fattura': invoice.get('invoice_number'),
            'Data': invoice.get('invoice_date'),
            'Fornitore': invoice.get('supplier_name'),
            'Importo': invoice.get('total_amount'),
            'Template': invoice.get('metadata_template', '')
        }
        
        # Aggiungi campi metadata
        for field_name, field_data in metadata.items():
            row[field_name] = field_data.get('value')
        
        rows.append(row)
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Fatture Metadata"
        
        if rows:
            # Header
            headers = list(rows[0].keys())
            ws.append(headers)
            
            # Stile header
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Dati
            for row in rows:
                ws.append(list(row.values()))
        
        # Salva
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        from fastapi.responses import StreamingResponse
        filename = f"metadata_report_{month_year or 'all'}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    except ImportError:
        raise ValidationError("openpyxl non installato")


# ==================== ENDPOINT 40: Statistiche Metadata ====================
@router.get(
    "/metadata/stats",
    summary="Statistiche metadata",
    description="Statistiche sull'utilizzo dei metadata"
)
async def get_metadata_stats(
    current_user: Dict[str, Any] = Depends(get_current_user)
) -> Dict[str, Any]:
    """
    Statistiche metadata.
    
    Mostra quante fatture hanno metadata,
    quali template sono più usati, ecc.
    """
    db = Database.get_db()
    invoices_collection = db[Collections.INVOICES]
    
    # Totale fatture
    total = await invoices_collection.count_documents({})
    
    # Fatture con metadata
    with_metadata = await invoices_collection.count_documents({
        'metadata': {'$exists': True, '$ne': {}}
    })
    
    # Template più usati
    pipeline = [
        {'$match': {'metadata_template': {'$exists': True, '$ne': None}}},
        {'$group': {'_id': '$metadata_template', 'count': {'$sum': 1}}},
        {'$sort': {'count': -1}},
        {'$limit': 10}
    ]
    
    cursor = invoices_collection.aggregate(pipeline)
    top_templates = await cursor.to_list(length=10)
    
    # Campi più usati
    all_invoices = await invoices_collection.find({
        'metadata': {'$exists': True, '$ne': {}}
    }).to_list(length=1000)
    
    field_usage = {}
    for invoice in all_invoices:
        metadata = invoice.get('metadata', {})
        for field_name in metadata.keys():
            field_usage[field_name] = field_usage.get(field_name, 0) + 1
    
    top_fields = sorted(field_usage.items(), key=lambda x: x[1], reverse=True)[:10]
    
    return {
        'total_invoices': total,
        'invoices_with_metadata': with_metadata,
        'percentage_with_metadata': round(with_metadata / total * 100, 2) if total > 0 else 0,
        'top_templates': [
            {'template_name': t['_id'], 'count': t['count']}
            for t in top_templates
        ],
        'top_fields': [
            {'field_name': f[0], 'usage_count': f[1]}
            for f in top_fields
        ]
    }


# ==================== ENDPOINT 41: Valida Metadata ====================
@router.post(
    "/metadata/validate",
    summary="Valida metadata",
    description="Valida che i metadata di una fattura siano completi e corretti"
)
async def validate_metadata(
    invoice_id: str = Query(..., description="ID fattura"),
    current_user: Dict[str, Any] = Depends(get_current_user)
) -> Dict[str, Any]:
    """
    Valida metadata.
    
    Verifica che tutti i campi obbligatori siano presenti
    e che i valori siano del tipo corretto.
    """
    db = Database.get_db()
    invoices_collection = db[Collections.INVOICES]
    
    invoice = await invoices_collection.find_one({'_id': invoice_id})
    if not invoice:
        raise NotFoundError(f"Invoice {invoice_id} not found")
    
    metadata = invoice.get('metadata', {})
    template_name = invoice.get('metadata_template')
    
    errors = []
    warnings = []
    
    # Se ha template, valida contro template
    if template_name:
        templates_collection = db[Collections.INVOICE_METADATA_TEMPLATES]
        template = await templates_collection.find_one({
            'template_name': template_name
        })
        
        if template:
            # Valida campi richiesti
            for field in template.get('fields', []):
                field_name = field['field_name']
                required = field.get('required', False)
                
                if required and field_name not in metadata:
                    errors.append(f"Campo obbligatorio mancante: {field_name}")
                
                if field_name in metadata:
                    # Valida tipo
                    expected_type = field['field_type']
                    actual_value = metadata[field_name].get('value')
                    
                    if expected_type == 'number' and not isinstance(actual_value, (int, float)):
                        errors.append(f"Campo {field_name} deve essere un numero")
                    elif expected_type == 'boolean' and not isinstance(actual_value, bool):
                        errors.append(f"Campo {field_name} deve essere boolean")
    else:
        warnings.append("Nessun template associato")
    
    # Controlla campi vuoti
    for field_name, field_data in metadata.items():
        if not field_data.get('value'):
            warnings.append(f"Campo {field_name} è vuoto")
    
    is_valid = len(errors) == 0
    
    return {
        'invoice_id': invoice_id,
        'invoice_number': invoice.get('invoice_number'),
        'valid': is_valid,
        'errors': errors,
        'warnings': warnings,
        'metadata_count': len(metadata),
        'template_name': template_name
    }


