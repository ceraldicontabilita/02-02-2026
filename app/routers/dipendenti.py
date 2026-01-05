"""
Gestione Dipendenti - Router API completo.
Anagrafica, turni, libro unico, libretti sanitari.
"""
from fastapi import APIRouter, HTTPException, Query, UploadFile, File, Body
from typing import Dict, Any, List, Optional
from datetime import datetime, date, timedelta
import uuid
import logging
import io
import re

from app.database import Database, Collections

logger = logging.getLogger(__name__)
router = APIRouter()

# Costanti
TURNI_TIPI = {
    "mattina": {"label": "Mattina", "orario": "06:00 - 14:00", "color": "#4caf50"},
    "pomeriggio": {"label": "Pomeriggio", "orario": "14:00 - 22:00", "color": "#2196f3"},
    "sera": {"label": "Sera", "orario": "18:00 - 02:00", "color": "#9c27b0"},
    "full": {"label": "Full Day", "orario": "10:00 - 22:00", "color": "#ff9800"},
    "riposo": {"label": "Riposo", "orario": "-", "color": "#9e9e9e"},
    "ferie": {"label": "Ferie", "orario": "-", "color": "#e91e63"},
    "malattia": {"label": "Malattia", "orario": "-", "color": "#f44336"}
}

MANSIONI = [
    "Cameriere", "Cuoco", "Aiuto Cuoco", "Barista", "Pizzaiolo", 
    "Lavapiatti", "Cassiera", "Responsabile Sala", "Chef", "Sommelier"
]

CONTRATTI_TIPI = [
    "Tempo Indeterminato", "Tempo Determinato", "Apprendistato", 
    "Stage/Tirocinio", "Collaborazione", "Part-time"
]


@router.get("")
async def list_dipendenti(
    skip: int = Query(0, ge=0),
    limit: int = Query(1000, ge=1, le=10000),
    attivo: Optional[bool] = Query(None),
    mansione: Optional[str] = Query(None),
    search: Optional[str] = Query(None)
) -> List[Dict[str, Any]]:
    """Lista dipendenti con filtri."""
    db = Database.get_db()
    
    query = {}
    if attivo is not None:
        query["attivo"] = attivo
    if mansione:
        query["mansione"] = mansione
    if search:
        query["$or"] = [
            {"nome_completo": {"$regex": search, "$options": "i"}},
            {"codice_fiscale": {"$regex": search, "$options": "i"}}
        ]
    
    dipendenti = await db[Collections.EMPLOYEES].find(query, {"_id": 0}).sort("nome_completo", 1).skip(skip).limit(limit).to_list(limit)
    return dipendenti


@router.get("/stats")
async def get_dipendenti_stats() -> Dict[str, Any]:
    """Statistiche dipendenti."""
    db = Database.get_db()
    
    total = await db[Collections.EMPLOYEES].count_documents({})
    attivi = await db[Collections.EMPLOYEES].count_documents({"attivo": {"$ne": False}})
    
    # Per mansione
    pipeline = [
        {"$group": {"_id": "$mansione", "count": {"$sum": 1}}}
    ]
    by_mansione = await db[Collections.EMPLOYEES].aggregate(pipeline).to_list(100)
    
    # Libretti in scadenza (prossimi 30 giorni)
    today = datetime.utcnow()
    deadline = today + timedelta(days=30)
    libretti_scadenza = await db[Collections.EMPLOYEES].count_documents({
        "libretto_scadenza": {"$lte": deadline.isoformat()[:10], "$gte": today.isoformat()[:10]}
    })
    
    return {
        "totale": total,
        "attivi": attivi,
        "inattivi": total - attivi,
        "per_mansione": {item["_id"] or "N/D": item["count"] for item in by_mansione},
        "libretti_in_scadenza": libretti_scadenza
    }


@router.get("/tipi-turno")
async def get_tipi_turno() -> Dict[str, Any]:
    """Ritorna i tipi di turno disponibili."""
    return TURNI_TIPI


@router.get("/mansioni")
async def get_mansioni() -> List[str]:
    """Ritorna le mansioni disponibili."""
    return MANSIONI


@router.get("/tipi-contratto")
async def get_tipi_contratto() -> List[str]:
    """Ritorna i tipi di contratto disponibili."""
    return CONTRATTI_TIPI


@router.post("")
async def create_dipendente(data: Dict[str, Any] = Body(...)) -> Dict[str, Any]:
    """Crea nuovo dipendente."""
    db = Database.get_db()
    
    # Campi obbligatori
    required = ["nome_completo"]
    for field in required:
        if not data.get(field):
            raise HTTPException(status_code=400, detail=f"Campo obbligatorio mancante: {field}")
    
    # Parse nome
    nome_parts = data["nome_completo"].split()
    cognome = nome_parts[0] if nome_parts else ""
    nome = " ".join(nome_parts[1:]) if len(nome_parts) > 1 else ""
    
    dipendente = {
        "id": str(uuid.uuid4()),
        "nome_completo": data["nome_completo"],
        "cognome": cognome,
        "nome": nome,
        "codice_fiscale": data.get("codice_fiscale", ""),
        "matricola": data.get("matricola", ""),
        "email": data.get("email", ""),
        "telefono": data.get("telefono", ""),
        "indirizzo": data.get("indirizzo", ""),
        "data_nascita": data.get("data_nascita"),
        "luogo_nascita": data.get("luogo_nascita", ""),
        "mansione": data.get("mansione", ""),
        "qualifica": data.get("qualifica", data.get("mansione", "")),
        "livello": data.get("livello", ""),
        "tipo_contratto": data.get("tipo_contratto", "Tempo Indeterminato"),
        "data_assunzione": data.get("data_assunzione"),
        "data_fine_contratto": data.get("data_fine_contratto"),
        "ore_settimanali": data.get("ore_settimanali", 40),
        "iban": data.get("iban", ""),
        # Libretto sanitario
        "libretto_numero": data.get("libretto_numero", ""),
        "libretto_scadenza": data.get("libretto_scadenza"),
        "libretto_file": data.get("libretto_file"),
        # Portale
        "portale_invitato": False,
        "portale_registrato": False,
        "portale_ultimo_accesso": None,
        # Status
        "attivo": True,
        "created_at": datetime.utcnow().isoformat(),
        "updated_at": datetime.utcnow().isoformat()
    }
    
    # Verifica duplicato CF
    if dipendente["codice_fiscale"]:
        existing = await db[Collections.EMPLOYEES].find_one({"codice_fiscale": dipendente["codice_fiscale"]})
        if existing:
            raise HTTPException(status_code=409, detail="Dipendente con questo codice fiscale già esistente")
    
    await db[Collections.EMPLOYEES].insert_one(dipendente)
    dipendente.pop("_id", None)
    
    return dipendente


# ============== BUSTE PAGA (must be before /{dipendente_id} to avoid route conflict) ==============

@router.get("/buste-paga")
async def get_buste_paga(
    anno: int = Query(...),
    mese: str = Query(...)
) -> List[Dict[str, Any]]:
    """
    Ottiene le buste paga per un determinato mese.
    Le buste paga vengono create automaticamente dai movimenti salari.
    """
    db = Database.get_db()
    
    periodo = f"{anno}-{mese}"
    
    # Cerca buste paga esistenti
    buste = await db["buste_paga"].find(
        {"periodo": periodo},
        {"_id": 0}
    ).to_list(1000)
    
    return buste


@router.post("/buste-paga")
async def create_busta_paga(data: Dict[str, Any] = Body(...)) -> Dict[str, Any]:
    """Crea o aggiorna una busta paga."""
    db = Database.get_db()
    
    required = ["dipendente_id", "periodo"]
    for field in required:
        if not data.get(field):
            raise HTTPException(status_code=400, detail=f"Campo {field} obbligatorio")
    
    # Cerca busta esistente
    existing = await db["buste_paga"].find_one({
        "dipendente_id": data["dipendente_id"],
        "periodo": data["periodo"]
    })
    
    busta = {
        "dipendente_id": data["dipendente_id"],
        "periodo": data["periodo"],
        "lordo": float(data.get("lordo", 0) or 0),
        "netto": float(data.get("netto", 0) or 0),
        "contributi": float(data.get("contributi", 0) or 0),
        "trattenute": float(data.get("trattenute", 0) or 0),
        "pagata": bool(data.get("pagata", False)),
        "data_pagamento": data.get("data_pagamento"),
        "note": data.get("note", ""),
        "updated_at": datetime.utcnow().isoformat()
    }
    
    if existing:
        await db["buste_paga"].update_one(
            {"id": existing["id"]},
            {"$set": busta}
        )
        busta["id"] = existing["id"]
    else:
        busta["id"] = str(uuid.uuid4())
        busta["created_at"] = datetime.utcnow().isoformat()
        await db["buste_paga"].insert_one(busta)
    
    busta.pop("_id", None)
    return busta


# ============== NOTA: Sezione SALARI rimossa ==============
# La gestione salari/prima nota è stata spostata in /app/app/routers/prima_nota_salari.py
# Endpoints disponibili: /api/prima-nota-salari/*

# ============== DIPENDENTE DETAIL (must be after specific routes) ==============

@router.get("/{dipendente_id}")
async def get_dipendente(dipendente_id: str) -> Dict[str, Any]:
    """Dettaglio singolo dipendente."""
    db = Database.get_db()
    
    dipendente = await db[Collections.EMPLOYEES].find_one(
        {"$or": [{"id": dipendente_id}, {"codice_fiscale": dipendente_id}]},
        {"_id": 0}
    )
    
    if not dipendente:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    return dipendente


@router.put("/{dipendente_id}")
async def update_dipendente(dipendente_id: str, data: Dict[str, Any] = Body(...)) -> Dict[str, str]:
    """Aggiorna dipendente."""
    db = Database.get_db()
    
    # Rimuovi campi non modificabili
    data.pop("id", None)
    data.pop("created_at", None)
    
    data["updated_at"] = datetime.utcnow().isoformat()
    
    result = await db[Collections.EMPLOYEES].update_one(
        {"$or": [{"id": dipendente_id}, {"codice_fiscale": dipendente_id}]},
        {"$set": data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    return {"message": "Dipendente aggiornato"}


@router.delete("/{dipendente_id}")
async def delete_dipendente(dipendente_id: str) -> Dict[str, str]:
    """Elimina dipendente."""
    db = Database.get_db()
    
    result = await db[Collections.EMPLOYEES].delete_one(
        {"$or": [{"id": dipendente_id}, {"codice_fiscale": dipendente_id}]}
    )
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    return {"message": "Dipendente eliminato"}


# ============== TURNI ==============

@router.get("/turni/settimana")
async def get_turni_settimana(
    data_inizio: str = Query(..., description="Data inizio settimana (YYYY-MM-DD)")
) -> Dict[str, Any]:
    """Ritorna i turni per una settimana."""
    db = Database.get_db()
    
    # Calcola date settimana
    start = datetime.strptime(data_inizio, "%Y-%m-%d")
    dates = [(start + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(7)]
    
    # Trova turni
    turni = await db["turni_dipendenti"].find(
        {"data": {"$in": dates}},
        {"_id": 0}
    ).to_list(1000)
    
    # Organizza per dipendente e data
    turni_by_employee = {}
    for t in turni:
        emp_id = t.get("dipendente_id")
        if emp_id not in turni_by_employee:
            turni_by_employee[emp_id] = {}
        turni_by_employee[emp_id][t.get("data")] = t.get("turno")
    
    # Carica dipendenti attivi
    dipendenti = await db[Collections.EMPLOYEES].find(
        {"attivo": {"$ne": False}},
        {"_id": 0, "id": 1, "nome_completo": 1, "mansione": 1}
    ).to_list(100)
    
    return {
        "settimana": dates,
        "dipendenti": dipendenti,
        "turni": turni_by_employee
    }


@router.post("/turni/salva")
async def salva_turni(data: Dict[str, Any] = Body(...)) -> Dict[str, str]:
    """Salva turni per una settimana."""
    db = Database.get_db()
    
    turni = data.get("turni", {})  # {dipendente_id: {data: turno}}
    
    for dip_id, turni_dip in turni.items():
        for data_turno, tipo_turno in turni_dip.items():
            await db["turni_dipendenti"].update_one(
                {"dipendente_id": dip_id, "data": data_turno},
                {"$set": {
                    "dipendente_id": dip_id,
                    "data": data_turno,
                    "turno": tipo_turno,
                    "updated_at": datetime.utcnow().isoformat()
                }},
                upsert=True
            )
    
    return {"message": "Turni salvati"}


# ============== LIBRETTI SANITARI ==============

@router.get("/libretti/scadenze")
async def get_libretti_scadenze(days: int = Query(30, ge=1, le=365)) -> List[Dict[str, Any]]:
    """Ritorna dipendenti con libretto in scadenza."""
    db = Database.get_db()
    
    today = datetime.utcnow()
    deadline = today + timedelta(days=days)
    
    dipendenti = await db[Collections.EMPLOYEES].find(
        {
            "libretto_scadenza": {"$ne": None},
            "$or": [
                {"libretto_scadenza": {"$lte": deadline.isoformat()[:10]}},
                {"libretto_scadenza": {"$lt": today.isoformat()[:10]}}  # Già scaduti
            ]
        },
        {"_id": 0}
    ).sort("libretto_scadenza", 1).to_list(100)
    
    return dipendenti


@router.put("/{dipendente_id}/libretto")
async def update_libretto(dipendente_id: str, data: Dict[str, Any] = Body(...)) -> Dict[str, str]:
    """Aggiorna dati libretto sanitario."""
    db = Database.get_db()
    
    update_data = {}
    if "libretto_numero" in data:
        update_data["libretto_numero"] = data["libretto_numero"]
    if "libretto_scadenza" in data:
        update_data["libretto_scadenza"] = data["libretto_scadenza"]
    if "libretto_file" in data:
        update_data["libretto_file"] = data["libretto_file"]
    
    update_data["updated_at"] = datetime.utcnow().isoformat()
    
    result = await db[Collections.EMPLOYEES].update_one(
        {"$or": [{"id": dipendente_id}, {"codice_fiscale": dipendente_id}]},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    return {"message": "Libretto aggiornato"}


# ============== PORTALE DIPENDENTI ==============

@router.post("/{dipendente_id}/invita-portale")
async def invita_portale(dipendente_id: str) -> Dict[str, str]:
    """Segna dipendente come invitato al portale."""
    db = Database.get_db()
    
    result = await db[Collections.EMPLOYEES].update_one(
        {"$or": [{"id": dipendente_id}, {"codice_fiscale": dipendente_id}]},
        {"$set": {
            "portale_invitato": True,
            "portale_data_invito": datetime.utcnow().isoformat(),
            "updated_at": datetime.utcnow().isoformat()
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    return {"message": "Invito inviato"}


@router.post("/invita-multipli")
async def invita_multipli(dipendenti_ids: List[str] = Body(...)) -> Dict[str, Any]:
    """Invita multipli dipendenti al portale."""
    db = Database.get_db()
    
    result = await db[Collections.EMPLOYEES].update_many(
        {"id": {"$in": dipendenti_ids}},
        {"$set": {
            "portale_invitato": True,
            "portale_data_invito": datetime.utcnow().isoformat(),
            "updated_at": datetime.utcnow().isoformat()
        }}
    )
    
    return {"message": f"Invitati {result.modified_count} dipendenti"}


# ============== LIBRETTI SANITARI - COLLECTION SEPARATA ==============

@router.get("/libretti-sanitari/all")
async def get_all_libretti_sanitari() -> List[Dict[str, Any]]:
    """Lista tutti i libretti sanitari."""
    db = Database.get_db()
    
    libretti = await db["libretti_sanitari"].find({}, {"_id": 0}).sort("data_scadenza", 1).to_list(500)
    return libretti


@router.post("/libretti-sanitari")
async def create_libretto_sanitario(data: Dict[str, Any] = Body(...)) -> Dict[str, Any]:
    """Crea nuovo libretto sanitario."""
    db = Database.get_db()
    
    libretto = {
        "id": str(uuid.uuid4()),
        "dipendente_nome": data.get("dipendente_nome", ""),
        "dipendente_id": data.get("dipendente_id"),
        "numero_libretto": data.get("numero_libretto", ""),
        "data_rilascio": data.get("data_rilascio"),
        "data_scadenza": data.get("data_scadenza"),
        "stato": data.get("stato", "valido"),
        "note": data.get("note", ""),
        "created_at": datetime.utcnow().isoformat(),
        "updated_at": datetime.utcnow().isoformat()
    }
    
    await db["libretti_sanitari"].insert_one(libretto)
    libretto.pop("_id", None)
    
    return libretto


@router.put("/libretti-sanitari/{libretto_id}")
async def update_libretto_sanitario(libretto_id: str, data: Dict[str, Any] = Body(...)) -> Dict[str, str]:
    """Aggiorna libretto sanitario."""
    db = Database.get_db()
    
    data.pop("id", None)
    data.pop("created_at", None)
    data["updated_at"] = datetime.utcnow().isoformat()
    
    result = await db["libretti_sanitari"].update_one(
        {"id": libretto_id},
        {"$set": data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Libretto non trovato")
    
    return {"message": "Libretto aggiornato"}


@router.delete("/libretti-sanitari/{libretto_id}")
async def delete_libretto_sanitario(libretto_id: str) -> Dict[str, str]:
    """Elimina libretto sanitario."""
    db = Database.get_db()
    
    result = await db["libretti_sanitari"].delete_one({"id": libretto_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Libretto non trovato")
    
    return {"message": "Libretto eliminato"}


# ============== LIBRO UNICO ==============

@router.get("/libro-unico/presenze")
async def get_libro_unico_presenze(month_year: str = Query(..., description="Formato: YYYY-MM")) -> List[Dict[str, Any]]:
    """Ottieni presenze dal libro unico per un mese."""
    db = Database.get_db()
    
    presenze = await db["libro_unico_presenze"].find(
        {"month_year": month_year},
        {"_id": 0}
    ).sort("dipendente_nome", 1).to_list(500)
    
    return presenze


@router.get("/libro-unico/salaries")
async def get_libro_unico_salaries(month_year: Optional[str] = Query(None, description="Formato: YYYY-MM")) -> List[Dict[str, Any]]:
    """Ottieni buste paga dal libro unico."""
    db = Database.get_db()
    
    query = {}
    if month_year:
        query["month_year"] = month_year
    
    salaries = await db["libro_unico_salaries"].find(query, {"_id": 0}).sort([("month_year", -1), ("dipendente_nome", 1)]).to_list(1000)
    
    return salaries


@router.post("/libro-unico/upload")
async def upload_libro_unico(
    file: UploadFile = File(...),
    month_year: str = Query(..., description="Formato: YYYY-MM")
) -> Dict[str, Any]:
    """
    Upload e parsing di PDF/Excel del libro unico.
    Estrae presenze, buste paga e aggiorna anagrafica.
    """
    db = Database.get_db()
    
    filename = file.filename.lower()
    content = await file.read()
    
    presenze_count = 0
    salaries_count = 0
    payments_count = 0
    anagrafica_created = 0
    anagrafica_updated = 0
    
    if filename.endswith('.pdf'):
        # Parsing PDF - estrai dati buste paga
        try:
            import fitz  # PyMuPDF
            
            pdf_doc = fitz.open(stream=content, filetype="pdf")
            
            for page in pdf_doc:
                text = page.get_text()
                
                # Cerca pattern busta paga
                # Pattern: NOME COGNOME seguito da valori numerici
                lines = text.split('\n')
                
                current_employee = None
                for i, line in enumerate(lines):
                    line = line.strip()
                    
                    # Cerca nome dipendente (tutte maiuscole, senza numeri)
                    if line and line.isupper() and not any(c.isdigit() for c in line) and len(line) > 3:
                        # Potrebbe essere un nome
                        if len(line.split()) >= 2:
                            current_employee = line
                    
                    # Cerca pattern "NETTO A PAGARE" o simili
                    if current_employee and 'netto' in line.lower():
                        # Cerca importo nelle righe successive
                        for j in range(i, min(i+5, len(lines))):
                            next_line = lines[j].strip()
                            # Cerca pattern €1.234,56 o 1234.56
                            import re
                            amounts = re.findall(r'[\d.,]+', next_line)
                            for amt in amounts:
                                try:
                                    # Converti formato italiano
                                    amt_clean = amt.replace('.', '').replace(',', '.')
                                    netto = float(amt_clean)
                                    if 100 < netto < 10000:  # Range plausibile per stipendio
                                        # Salva busta paga
                                        salary_doc = {
                                            "id": str(uuid.uuid4()),
                                            "dipendente_nome": current_employee,
                                            "month_year": month_year,
                                            "netto_a_pagare": netto,
                                            "acconto_pagato": 0,
                                            "differenza": netto,
                                            "note": f"Importato da PDF: {file.filename}",
                                            "created_at": datetime.utcnow().isoformat()
                                        }
                                        
                                        # Verifica se esiste già
                                        existing = await db["libro_unico_salaries"].find_one({
                                            "dipendente_nome": current_employee,
                                            "month_year": month_year
                                        })
                                        
                                        if not existing:
                                            await db["libro_unico_salaries"].insert_one(salary_doc)
                                            salaries_count += 1
                                        
                                        current_employee = None
                                        break
                                except:
                                    continue
                            if current_employee is None:
                                break
            
            pdf_doc.close()
            
        except ImportError:
            raise HTTPException(status_code=500, detail="PyMuPDF non installato per parsing PDF")
        except Exception as e:
            logger.error(f"Errore parsing PDF: {e}")
            raise HTTPException(status_code=400, detail=f"Errore parsing PDF: {str(e)}")
    
    elif filename.endswith(('.xlsx', '.xls')):
        # Parsing Excel
        try:
            import pandas as pd
            
            df = pd.read_excel(io.BytesIO(content))
            
            # Cerca colonne rilevanti
            columns_lower = {c.lower(): c for c in df.columns}
            
            for idx, row in df.iterrows():
                try:
                    # Cerca nome dipendente
                    nome = None
                    for col_key in ['nome', 'dipendente', 'cognome e nome', 'nominativo']:
                        if col_key in columns_lower:
                            nome = str(row[columns_lower[col_key]]) if pd.notna(row[columns_lower[col_key]]) else None
                            break
                    
                    if not nome or nome == 'nan':
                        continue
                    
                    # Cerca importo netto
                    netto = None
                    for col_key in ['netto', 'netto a pagare', 'importo', 'stipendio']:
                        if col_key in columns_lower:
                            val = row[columns_lower[col_key]]
                            if pd.notna(val):
                                try:
                                    netto = float(val)
                                except:
                                    pass
                            break
                    
                    if netto and netto > 0:
                        salary_doc = {
                            "id": str(uuid.uuid4()),
                            "dipendente_nome": nome.upper(),
                            "month_year": month_year,
                            "netto_a_pagare": netto,
                            "acconto_pagato": 0,
                            "differenza": netto,
                            "note": f"Importato da Excel: {file.filename}",
                            "created_at": datetime.utcnow().isoformat()
                        }
                        
                        existing = await db["libro_unico_salaries"].find_one({
                            "dipendente_nome": nome.upper(),
                            "month_year": month_year
                        })
                        
                        if not existing:
                            await db["libro_unico_salaries"].insert_one(salary_doc)
                            salaries_count += 1
                        
                except Exception as e:
                    logger.warning(f"Errore riga {idx}: {e}")
                    continue
                    
        except Exception as e:
            logger.error(f"Errore parsing Excel: {e}")
            raise HTTPException(status_code=400, detail=f"Errore parsing Excel: {str(e)}")
    
    else:
        raise HTTPException(status_code=400, detail="Formato non supportato. Usa PDF o Excel.")
    
    return {
        "success": True,
        "message": f"Import completato per {month_year}",
        "presenze_count": presenze_count,
        "salaries_count": salaries_count,
        "payments_count": payments_count,
        "anagrafica_created": anagrafica_created,
        "anagrafica_updated": anagrafica_updated
    }


@router.get("/libro-unico/export-excel")
async def export_libro_unico_excel(month_year: str = Query(..., description="Formato: YYYY-MM")):
    """Esporta libro unico in Excel."""
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    db = Database.get_db()
    
    # Recupera dati
    salaries = await db["libro_unico_salaries"].find(
        {"month_year": month_year},
        {"_id": 0}
    ).sort("dipendente_nome", 1).to_list(500)
    
    # Crea workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Libro Unico {month_year}"
    
    # Header
    headers = ["Dipendente", "Netto a Pagare", "Acconto", "Differenza", "Note"]
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Dati
    for row_num, salary in enumerate(salaries, 2):
        ws.cell(row=row_num, column=1, value=salary.get("dipendente_nome", ""))
        ws.cell(row=row_num, column=2, value=salary.get("netto_a_pagare", 0))
        ws.cell(row=row_num, column=3, value=salary.get("acconto_pagato", 0))
        ws.cell(row=row_num, column=4, value=salary.get("differenza", 0))
        ws.cell(row=row_num, column=5, value=salary.get("note", ""))
    
    # Larghezze colonne
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    
    # Salva
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=libro_unico_{month_year}.xlsx"}
    )


@router.put("/libro-unico/salaries/{salary_id}")
async def update_libro_unico_salary(
    salary_id: str,
    netto_a_pagare: float = Query(...),
    acconto_pagato: float = Query(0),
    differenza: float = Query(...),
    note: str = Query("")
) -> Dict[str, str]:
    """Aggiorna busta paga libro unico."""
    db = Database.get_db()
    
    result = await db["libro_unico_salaries"].update_one(
        {"id": salary_id},
        {"$set": {
            "netto_a_pagare": netto_a_pagare,
            "acconto_pagato": acconto_pagato,
            "differenza": differenza,
            "note": note,
            "updated_at": datetime.utcnow().isoformat()
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Record non trovato")
    
    return {"message": "Aggiornato"}


@router.delete("/libro-unico/salaries/{salary_id}")
async def delete_libro_unico_salary(salary_id: str) -> Dict[str, str]:
    """Elimina busta paga libro unico."""
    db = Database.get_db()
    
    result = await db["libro_unico_salaries"].delete_one({"id": salary_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Record non trovato")
    
    return {"message": "Eliminato"}


# Note: salari and buste-paga routes are defined earlier in the file to avoid route conflict with /{dipendente_id}



@router.get("/portale/stats")
async def get_portale_stats() -> Dict[str, Any]:
    """Statistiche portale dipendenti."""
    db = Database.get_db()
    
    total = await db[Collections.EMPLOYEES].count_documents({"attivo": {"$ne": False}})
    invitati = await db[Collections.EMPLOYEES].count_documents({"portale_invitato": True})
    registrati = await db[Collections.EMPLOYEES].count_documents({"portale_registrato": True})
    mai_invitati = await db[Collections.EMPLOYEES].count_documents({
        "attivo": {"$ne": False},
        "$or": [{"portale_invitato": False}, {"portale_invitato": {"$exists": False}}]
    })
    
    return {
        "totale": total,
        "mai_invitati": mai_invitati,
        "invitati": invitati,
        "registrati": registrati
    }


