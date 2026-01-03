# IVA ANNUAL REPORT API - DA INTEGRARE IN server.py

# ==================== IVA ANNUALE ====================

@api_router.get("/iva/annual-report/{year}")
async def get_iva_annual_report(year: int, username: str = Depends(get_current_user)):
    """
    Riepilogo IVA annuale con calcolo mensile.
    
    NORMATIVA:
    - IVA CREDITO: Da fatture passive (acquisti) ricevute e registrate
    - IVA DEBITO: Da corrispettivi e fatture attive
    - REGOLA 15 GIORNI: Fatture ricevute entro il 15 del mese successivo possono essere 
      registrate nella liquidazione del mese precedente (competenza)
    
    Returns: Tabella mensile con IVA credito, debito e saldo per ogni mese
    """
    try:
        monthly_data = []
        
        for month in range(1, 13):
            month_str = f"{year}-{month:02d}"
            
            # Calcola IVA CREDITO da fatture passive
            # Include fatture con data fino al 15 del mese successivo se doc_date è nel mese corrente
            next_month = month + 1 if month < 12 else 1
            next_year = year if month < 12 else year + 1
            limit_date = f"{next_year}-{next_month:02d}-15"
            
            # Query per IVA credito
            iva_credito_pipeline = [
                {
                    "$match": {
                        "user_id": username,
                        "$or": [
                            # Fatture con data nel mese corrente
                            {"invoice_date": {"$regex": f"^{month_str}"}},
                            # Fatture ricevute entro il 15 del mese successivo con data documento nel mese
                            {
                                "$and": [
                                    {"invoice_date": {"$regex": f"^{month_str}"}},
                                    {"created_at": {"$lte": limit_date}}
                                ]
                            }
                        ]
                    }
                },
                {
                    "$group": {
                        "_id": None,
                        "total_iva": {"$sum": "$iva_amount"},
                        "count": {"$sum": 1}
                    }
                }
            ]
            
            iva_credito_result = await db.invoices.aggregate(iva_credito_pipeline).to_list(1)
            iva_credito = iva_credito_result[0]['total_iva'] if iva_credito_result else 0
            fatture_passive_count = iva_credito_result[0]['count'] if iva_credito_result else 0
            
            # Calcola IVA DEBITO da corrispettivi
            # I corrispettivi sono registrati alla data dell'operazione
            iva_debito_pipeline = [
                {
                    "$match": {
                        "user_id": username,
                        "date": {"$regex": f"^{month_str}"}
                    }
                },
                {
                    "$group": {
                        "_id": None,
                        "total_iva": {"$sum": "$iva"},
                        "count": {"$sum": 1}
                    }
                }
            ]
            
            iva_debito_result = await db.cash_movements.aggregate(iva_debito_pipeline).to_list(1)
            iva_debito = iva_debito_result[0]['total_iva'] if iva_debito_result else 0
            corrispettivi_count = iva_debito_result[0]['count'] if iva_debito_result else 0
            
            # Calcola saldo (negativo = da versare, positivo = a credito)
            saldo = iva_debito - iva_credito
            
            monthly_data.append({
                "mese": month,
                "mese_nome": datetime(year, month, 1).strftime("%B"),
                "anno": year,
                "iva_credito": round(iva_credito, 2),
                "iva_debito": round(iva_debito, 2),
                "saldo": round(saldo, 2),
                "stato": "Da versare" if saldo > 0 else "A credito" if saldo < 0 else "Pareggio",
                "fatture_passive": fatture_passive_count,
                "corrispettivi": corrispettivi_count
            })
        
        # Calcola totali annuali
        total_iva_credito = sum(m['iva_credito'] for m in monthly_data)
        total_iva_debito = sum(m['iva_debito'] for m in monthly_data)
        total_saldo = total_iva_debito - total_iva_credito
        
        return {
            "anno": year,
            "monthly_data": monthly_data,
            "totali": {
                "iva_credito": round(total_iva_credito, 2),
                "iva_debito": round(total_iva_debito, 2),
                "saldo": round(total_saldo, 2),
                "stato": "Da versare" if total_saldo > 0 else "A credito" if total_saldo < 0 else "Pareggio"
            }
        }
        
    except Exception as e:
        logger.error(f"❌ Errore calcolo IVA annuale: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/iva/month-details/{year}/{month}")
async def get_iva_month_details(year: int, month: int, username: str = Depends(get_current_user)):
    """Get detailed IVA data for a specific month"""
    try:
        month_str = f"{year}-{month:02d}"
        
        # Fatture passive (IVA credito)
        fatture_passive = await db.invoices.find({
            "user_id": username,
            "invoice_date": {"$regex": f"^{month_str}"}
        }, {"_id": 0}).sort("invoice_date", 1).to_list(1000)
        
        # Corrispettivi (IVA debito)
        corrispettivi = await db.cash_movements.find({
            "user_id": username,
            "date": {"$regex": f"^{month_str}"}
        }, {"_id": 0}).sort("date", 1).to_list(1000)
        
        return {
            "anno": year,
            "mese": month,
            "fatture_passive": fatture_passive,
            "corrispettivi": corrispettivi,
            "totale_fatture": len(fatture_passive),
            "totale_corrispettivi": len(corrispettivi)
        }
        
    except Exception as e:
        logger.error(f"❌ Errore dettagli IVA mese: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/iva/export-annual-excel/{year}")
async def export_iva_annual_excel(year: int, username: str = Depends(get_current_user)):
    """Export IVA annual report to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from io import BytesIO
        
        # Get data
        report_response = await get_iva_annual_report(year, username)
        monthly_data = report_response['monthly_data']
        totali = report_response['totali']
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"IVA {year}"
        
        # Header styling
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = f"RIEPILOGO IVA ANNUALE {year}"
        title_cell.font = Font(bold=True, size=16, color="0066CC")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Headers
        headers = ['Mese', 'IVA Credito', 'IVA Debito', 'Saldo', 'Stato', 'N° Fatture', 'N° Corrispettivi']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data rows
        row_num = 4
        for month_data in monthly_data:
            ws.cell(row=row_num, column=1, value=month_data['mese_nome'])
            
            credito_cell = ws.cell(row=row_num, column=2, value=month_data['iva_credito'])
            credito_cell.number_format = '€#,##0.00'
            
            debito_cell = ws.cell(row=row_num, column=3, value=month_data['iva_debito'])
            debito_cell.number_format = '€#,##0.00'
            
            saldo_cell = ws.cell(row=row_num, column=4, value=month_data['saldo'])
            saldo_cell.number_format = '€#,##0.00'
            if month_data['saldo'] > 0:
                saldo_cell.font = Font(color="FF0000")  # Rosso per da versare
            elif month_data['saldo'] < 0:
                saldo_cell.font = Font(color="00FF00")  # Verde per credito
            
            ws.cell(row=row_num, column=5, value=month_data['stato'])
            ws.cell(row=row_num, column=6, value=month_data['fatture_passive'])
            ws.cell(row=row_num, column=7, value=month_data['corrispettivi'])
            
            # Apply border
            for col_num in range(1, 8):
                ws.cell(row=row_num, column=col_num).border = border
            
            row_num += 1
        
        # Total row
        ws.cell(row=row_num, column=1, value="TOTALE ANNO").font = Font(bold=True)
        
        total_credito = ws.cell(row=row_num, column=2, value=totali['iva_credito'])
        total_credito.number_format = '€#,##0.00'
        total_credito.font = Font(bold=True)
        total_credito.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        total_debito = ws.cell(row=row_num, column=3, value=totali['iva_debito'])
        total_debito.number_format = '€#,##0.00'
        total_debito.font = Font(bold=True)
        total_debito.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        total_saldo = ws.cell(row=row_num, column=4, value=totali['saldo'])
        total_saldo.number_format = '€#,##0.00'
        total_saldo.font = Font(bold=True, size=12)
        total_saldo.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        
        ws.cell(row=row_num, column=5, value=totali['stato']).font = Font(bold=True)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 18
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        from fastapi.responses import StreamingResponse
        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=riepilogo_iva_{year}.xlsx"
            }
        )
        
    except Exception as e:
        logger.error(f"❌ Errore export IVA Excel: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
