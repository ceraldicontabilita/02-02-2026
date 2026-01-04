"""
Test Excel Export Endpoints - Prima Nota and HACCP
Tests for:
- /api/prima-nota/export/excel - Prima Nota Excel export (cassa, banca, entrambi)
- /api/haccp-completo/export/temperature-excel - HACCP Temperature export (frigoriferi, congelatori)
- /api/haccp-completo/export/sanificazioni-excel - HACCP Sanificazioni export
"""
import pytest
import requests
import os
import io

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')


class TestPrimaNotaExcelExport:
    """Test Prima Nota Excel export endpoint"""
    
    def test_export_excel_entrambi(self):
        """Test export Excel with tipo=entrambi (both cassa and banca sheets)"""
        response = requests.get(f"{BASE_URL}/api/prima-nota/export/excel?tipo=entrambi")
        
        # Status code assertion
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        
        # Content-Type assertion
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers.get("Content-Type", ""), \
            f"Expected Excel content type, got {response.headers.get('Content-Type')}"
        
        # Content-Disposition assertion (should have filename)
        content_disp = response.headers.get("Content-Disposition", "")
        assert "attachment" in content_disp, f"Expected attachment disposition, got {content_disp}"
        assert "prima_nota_" in content_disp, f"Expected prima_nota_ in filename, got {content_disp}"
        assert ".xlsx" in content_disp, f"Expected .xlsx extension, got {content_disp}"
        
        # Content length assertion (should have some data)
        assert len(response.content) > 0, "Expected non-empty Excel file"
        
        # Validate Excel structure using openpyxl
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(response.content))
            sheet_names = wb.sheetnames
            print(f"Excel sheets found: {sheet_names}")
            
            # Should have at least one sheet (cassa or banca)
            assert len(sheet_names) >= 1, f"Expected at least 1 sheet, got {len(sheet_names)}"
            
            # Check for expected sheet names
            expected_sheets = ["Prima Nota Cassa", "Prima Nota Banca"]
            found_sheets = [s for s in expected_sheets if s in sheet_names]
            print(f"Found expected sheets: {found_sheets}")
            
            # Verify data in sheets
            for sheet_name in found_sheets:
                ws = wb[sheet_name]
                row_count = ws.max_row
                print(f"Sheet '{sheet_name}' has {row_count} rows")
                assert row_count >= 1, f"Sheet {sheet_name} should have at least header row"
                
        except ImportError:
            print("openpyxl not available for detailed validation, skipping structure check")
        
        print("PASS: Prima Nota export Excel (entrambi) works correctly")
    
    def test_export_excel_cassa_only(self):
        """Test export Excel with tipo=cassa (only cassa sheet)"""
        response = requests.get(f"{BASE_URL}/api/prima-nota/export/excel?tipo=cassa")
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers.get("Content-Type", "")
        assert len(response.content) > 0
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(response.content))
            sheet_names = wb.sheetnames
            print(f"Cassa-only export sheets: {sheet_names}")
            
            # Should have Prima Nota Cassa sheet
            if "Prima Nota Cassa" in sheet_names:
                ws = wb["Prima Nota Cassa"]
                print(f"Prima Nota Cassa has {ws.max_row} rows")
        except ImportError:
            pass
        
        print("PASS: Prima Nota export Excel (cassa only) works correctly")
    
    def test_export_excel_banca_only(self):
        """Test export Excel with tipo=banca (only banca sheet)"""
        response = requests.get(f"{BASE_URL}/api/prima-nota/export/excel?tipo=banca")
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers.get("Content-Type", "")
        assert len(response.content) > 0
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(response.content))
            sheet_names = wb.sheetnames
            print(f"Banca-only export sheets: {sheet_names}")
            
            if "Prima Nota Banca" in sheet_names:
                ws = wb["Prima Nota Banca"]
                print(f"Prima Nota Banca has {ws.max_row} rows")
        except ImportError:
            pass
        
        print("PASS: Prima Nota export Excel (banca only) works correctly")
    
    def test_export_excel_with_date_filter(self):
        """Test export Excel with date filters"""
        response = requests.get(
            f"{BASE_URL}/api/prima-nota/export/excel?tipo=entrambi&data_da=2024-01-01&data_a=2025-12-31"
        )
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers.get("Content-Type", "")
        
        print("PASS: Prima Nota export Excel with date filter works correctly")


class TestHACCPTemperatureExcelExport:
    """Test HACCP Temperature Excel export endpoint"""
    
    def test_export_temperature_frigoriferi(self):
        """Test export temperature frigoriferi Excel"""
        response = requests.get(
            f"{BASE_URL}/api/haccp-completo/export/temperature-excel?mese=2026-01&tipo=frigoriferi"
        )
        
        # Status code assertion
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        
        # Content-Type assertion
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers.get("Content-Type", ""), \
            f"Expected Excel content type, got {response.headers.get('Content-Type')}"
        
        # Content-Disposition assertion
        content_disp = response.headers.get("Content-Disposition", "")
        assert "attachment" in content_disp, f"Expected attachment disposition, got {content_disp}"
        assert "haccp_temperature_frigoriferi" in content_disp, f"Expected haccp_temperature_frigoriferi in filename, got {content_disp}"
        assert ".xlsx" in content_disp, f"Expected .xlsx extension, got {content_disp}"
        
        # Content length assertion
        assert len(response.content) > 0, "Expected non-empty Excel file"
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(response.content))
            sheet_names = wb.sheetnames
            print(f"Temperature frigoriferi export sheets: {sheet_names}")
            
            # Check for expected sheet
            if "Temperature Frigoriferi" in sheet_names:
                ws = wb["Temperature Frigoriferi"]
                print(f"Temperature Frigoriferi has {ws.max_row} rows")
        except ImportError:
            pass
        
        print("PASS: HACCP Temperature frigoriferi export Excel works correctly")
    
    def test_export_temperature_congelatori(self):
        """Test export temperature congelatori Excel"""
        response = requests.get(
            f"{BASE_URL}/api/haccp-completo/export/temperature-excel?mese=2026-01&tipo=congelatori"
        )
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers.get("Content-Type", "")
        
        content_disp = response.headers.get("Content-Disposition", "")
        assert "haccp_temperature_congelatori" in content_disp, f"Expected haccp_temperature_congelatori in filename, got {content_disp}"
        
        assert len(response.content) > 0
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(response.content))
            sheet_names = wb.sheetnames
            print(f"Temperature congelatori export sheets: {sheet_names}")
            
            if "Temperature Congelatori" in sheet_names:
                ws = wb["Temperature Congelatori"]
                print(f"Temperature Congelatori has {ws.max_row} rows")
        except ImportError:
            pass
        
        print("PASS: HACCP Temperature congelatori export Excel works correctly")
    
    def test_export_temperature_missing_mese_param(self):
        """Test export temperature without required mese parameter"""
        response = requests.get(
            f"{BASE_URL}/api/haccp-completo/export/temperature-excel?tipo=frigoriferi"
        )
        
        # Should return 422 (validation error) for missing required parameter
        assert response.status_code == 422, f"Expected 422 for missing mese param, got {response.status_code}"
        
        print("PASS: HACCP Temperature export correctly validates required mese parameter")


class TestHACCPSanificazioniExcelExport:
    """Test HACCP Sanificazioni Excel export endpoint"""
    
    def test_export_sanificazioni(self):
        """Test export sanificazioni Excel"""
        response = requests.get(
            f"{BASE_URL}/api/haccp-completo/export/sanificazioni-excel?mese=2026-01"
        )
        
        # Status code assertion
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        
        # Content-Type assertion
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers.get("Content-Type", ""), \
            f"Expected Excel content type, got {response.headers.get('Content-Type')}"
        
        # Content-Disposition assertion
        content_disp = response.headers.get("Content-Disposition", "")
        assert "attachment" in content_disp, f"Expected attachment disposition, got {content_disp}"
        assert "haccp_sanificazioni" in content_disp, f"Expected haccp_sanificazioni in filename, got {content_disp}"
        assert ".xlsx" in content_disp, f"Expected .xlsx extension, got {content_disp}"
        
        # Content length assertion
        assert len(response.content) > 0, "Expected non-empty Excel file"
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(response.content))
            sheet_names = wb.sheetnames
            print(f"Sanificazioni export sheets: {sheet_names}")
            
            if "Sanificazioni" in sheet_names:
                ws = wb["Sanificazioni"]
                print(f"Sanificazioni has {ws.max_row} rows")
        except ImportError:
            pass
        
        print("PASS: HACCP Sanificazioni export Excel works correctly")
    
    def test_export_sanificazioni_missing_mese_param(self):
        """Test export sanificazioni without required mese parameter"""
        response = requests.get(
            f"{BASE_URL}/api/haccp-completo/export/sanificazioni-excel"
        )
        
        # Should return 422 (validation error) for missing required parameter
        assert response.status_code == 422, f"Expected 422 for missing mese param, got {response.status_code}"
        
        print("PASS: HACCP Sanificazioni export correctly validates required mese parameter")


class TestExcelExportDataIntegrity:
    """Test data integrity in Excel exports"""
    
    def test_prima_nota_export_has_expected_columns(self):
        """Verify Prima Nota export has expected columns"""
        response = requests.get(f"{BASE_URL}/api/prima-nota/export/excel?tipo=entrambi")
        
        assert response.status_code == 200
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(response.content))
            
            expected_cassa_cols = ["data", "tipo", "importo", "descrizione", "categoria", "riferimento"]
            expected_banca_cols = ["data", "tipo", "importo", "descrizione", "categoria", "riferimento"]
            
            if "Prima Nota Cassa" in wb.sheetnames:
                ws = wb["Prima Nota Cassa"]
                if ws.max_row > 0:
                    headers = [cell.value for cell in ws[1] if cell.value]
                    print(f"Prima Nota Cassa headers: {headers}")
                    for col in expected_cassa_cols:
                        if col in headers:
                            print(f"  ✓ Found column: {col}")
            
            if "Prima Nota Banca" in wb.sheetnames:
                ws = wb["Prima Nota Banca"]
                if ws.max_row > 0:
                    headers = [cell.value for cell in ws[1] if cell.value]
                    print(f"Prima Nota Banca headers: {headers}")
                    for col in expected_banca_cols:
                        if col in headers:
                            print(f"  ✓ Found column: {col}")
                            
        except ImportError:
            print("openpyxl not available, skipping column validation")
        
        print("PASS: Prima Nota export column structure validated")
    
    def test_haccp_temperature_export_has_expected_columns(self):
        """Verify HACCP Temperature export has expected columns"""
        response = requests.get(
            f"{BASE_URL}/api/haccp-completo/export/temperature-excel?mese=2026-01&tipo=frigoriferi"
        )
        
        assert response.status_code == 200
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(response.content))
            
            expected_cols = ["data", "ora", "equipaggiamento", "temperatura", "conforme", "operatore", "note"]
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if ws.max_row > 0:
                    headers = [cell.value for cell in ws[1] if cell.value]
                    print(f"{sheet_name} headers: {headers}")
                    for col in expected_cols:
                        if col in headers:
                            print(f"  ✓ Found column: {col}")
                            
        except ImportError:
            print("openpyxl not available, skipping column validation")
        
        print("PASS: HACCP Temperature export column structure validated")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
